#!/usr/bin/env python3
from __future__ import annotations

import argparse
from contextlib import ExitStack
from datetime import datetime, timedelta
import importlib.util
import json
import os
from pathlib import Path
import re
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from vendor.config import ODPSConfig
from vendor.odps_connector import ODPSConnector

EXCEL_MAX_ROWS_PER_SHEET = int(os.environ.get("ODPS_EXPORT_MAX_ROWS_PER_SHEET", "1048576"))
PARQUET_BATCH_SIZE = int(os.environ.get("ODPS_EXPORT_PARQUET_BATCH_SIZE", "20000"))
EXCEL_BATCH_SIZE = int(os.environ.get("ODPS_EXPORT_EXCEL_BATCH_SIZE", "20000"))
TEMPLATE_PATTERN = re.compile(r"\$\{([A-Za-z_][A-Za-z0-9_]*)\}")


def _load_config(config_path: Path) -> list[dict[str, Any]]:
    spec = importlib.util.spec_from_file_location("odps_export_config", config_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load config: {config_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    exports = getattr(module, "EXPORTS", None)
    if not isinstance(exports, list) or not exports:
        raise ValueError("EXPORTS must be a non-empty list")
    return exports


def _default_params() -> dict[str, str]:
    now = datetime.now()
    today = now.strftime("%Y%m%d")
    yesterday = (now - timedelta(days=1)).strftime("%Y%m%d")
    return {
        "today": today,
        "yesterday": yesterday,
    }


def _parse_cli_params(raw_params: list[str]) -> dict[str, str]:
    params: dict[str, str] = {}
    for raw in raw_params:
        if "=" not in raw:
            raise ValueError(f"invalid --param '{raw}', expected KEY=VALUE")
        key, value = raw.split("=", 1)
        key = key.strip()
        if not key:
            raise ValueError(f"invalid --param '{raw}', empty key")
        params[key] = value
    return params


def _render_template(value: str, params: dict[str, Any]) -> str:
    def repl(match: re.Match[str]) -> str:
        key = match.group(1)
        if key not in params:
            raise KeyError(f"missing template param: {key}")
        return str(params[key])

    return TEMPLATE_PATTERN.sub(repl, value)


def _build_params(spec: dict[str, Any], cli_params: dict[str, str]) -> dict[str, Any]:
    params: dict[str, Any] = {}
    params.update(_default_params())
    params.update(cli_params)
    raw_spec_params = spec.get("params") or {}
    if not isinstance(raw_spec_params, dict):
        raise ValueError("spec params must be a dict when provided")

    pending = {str(key): str(value) for key, value in raw_spec_params.items()}
    for _ in range(len(pending) + 1):
        changed = False
        for key, value in pending.items():
            rendered = _render_template(value, params)
            if params.get(key) != rendered:
                params[key] = rendered
                changed = True
        if not changed:
            break
    return params


def _resolve_string(value: str | None, params: dict[str, Any]) -> str | None:
    if value is None:
        return None
    return _render_template(str(value), params)


def _build_sql(spec: dict[str, Any], config_dir: Path, params: dict[str, Any]) -> str:
    sql_file = spec.get("sql_file")
    if sql_file:
        sql_path = Path(_resolve_string(str(sql_file), params) or "")
        if not sql_path.is_absolute():
            sql_path = config_dir / sql_path
        sql = sql_path.read_text(encoding="utf-8")
        return _render_template(sql.strip().rstrip(";"), params)

    sql = spec.get("sql")
    if sql:
        return _render_template(str(sql).strip().rstrip(";"), params)
    raise ValueError("each export must define sql or sql_file")


def _header_row(ws, headers: list[str]) -> None:
    cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = Font(bold=True)
        cells.append(cell)
    ws.append(cells)


def _normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _stage_parquet_path(output: Path, spec: dict[str, Any]) -> Path:
    if parquet_output := spec.get("parquet_output"):
        return Path(parquet_output)
    if output.suffix.lower() == ".parquet":
        return output
    return output.with_suffix(".parquet")


def _parquet_to_xlsx(parquet_path: Path, output: Path, base_sheet_name: str, rename_map: dict[str, str]) -> int:
    parquet_file = pq.ParquetFile(parquet_path)
    workbook = Workbook(write_only=True)
    total_rows = 0
    sheet_index = 1
    ws = None
    rows_in_sheet = 0
    for batch in parquet_file.iter_batches(batch_size=EXCEL_BATCH_SIZE):
        batch_table = pa.Table.from_batches([batch])
        columns = batch_table.column_names
        headers = [rename_map.get(col, col) for col in columns]
        if ws is None:
            ws = workbook.create_sheet(title=(base_sheet_name or "Sheet1")[:31])
            _header_row(ws, headers)
            rows_in_sheet = 1
        column_values = [batch_table.column(col).to_pylist() for col in columns]
        for row in zip(*column_values):
            if rows_in_sheet >= EXCEL_MAX_ROWS_PER_SHEET:
                sheet_index += 1
                title = f"{base_sheet_name}_{sheet_index}"[:31]
                ws = workbook.create_sheet(title=title)
                _header_row(ws, headers)
                rows_in_sheet = 1
            ws.append(row)
            total_rows += 1
            rows_in_sheet += 1
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return total_rows


def _fetch_query_to_parquet(sql: str, output: Path) -> tuple[int, list[str]]:
    output.parent.mkdir(parents=True, exist_ok=True)
    with ExitStack() as stack:
        connector = ODPSConnector(ODPSConfig())
        stack.callback(connector.close)
        odps_client = connector.connect()
        instance = odps_client.execute_sql(sql)
        reader = stack.enter_context(instance.open_reader())

        columns = [col.name for col in reader._schema.columns]
        batch_columns = {col: [] for col in columns}
        total_rows = 0
        schema = pa.schema([pa.field(col, pa.string()) for col in columns])

        def make_table(column_batches: dict[str, list[str]]) -> pa.Table:
            return pa.Table.from_pydict(
                {
                    col: pa.array(column_batches[col], type=pa.string())
                    for col in columns
                },
                schema=schema,
            )

        def append_record(record: Any) -> None:
            if isinstance(record, dict):
                for col in columns:
                    batch_columns[col].append(_normalize_value(record.get(col)))
                return
            for col in columns:
                batch_columns[col].append(_normalize_value(record[col]))

        def flush_batch(writer: pq.ParquetWriter | None) -> pq.ParquetWriter | None:
            if not batch_columns[columns[0]]:
                return writer
            table = make_table(batch_columns)
            if writer is None:
                writer = stack.enter_context(pq.ParquetWriter(output, schema))
            writer.write_table(table)
            for col in columns:
                batch_columns[col] = []
            return writer

        reader_iter = iter(reader)
        first_row = next(reader_iter, None)
        if first_row is None:
            empty_table = pa.Table.from_pydict(
                {col: pa.array([], type=pa.string()) for col in columns},
                schema=schema,
            )
            with pq.ParquetWriter(output, empty_table.schema) as writer:
                writer.write_table(empty_table)
            return 0, columns

        append_record(first_row)
        total_rows += 1

        writer = None
        for record in reader_iter:
            append_record(record)
            total_rows += 1
            if len(batch_columns[columns[0]]) >= PARQUET_BATCH_SIZE:
                writer = flush_batch(writer)

        writer = flush_batch(writer)

        return total_rows, columns


def _select_exports(exports: list[dict[str, Any]], selected_names: list[str]) -> list[dict[str, Any]]:
    if not selected_names:
        return exports

    selected = []
    seen = set()
    wanted = {name for name in selected_names if name}
    for spec in exports:
        name = str(spec.get("name") or "")
        if name in wanted:
            selected.append(spec)
            seen.add(name)

    missing = sorted(wanted - seen)
    if missing:
        raise ValueError(f"unknown export name(s): {', '.join(missing)}")
    return selected


def _run_export(
    spec: dict[str, Any],
    config_dir: Path,
    cli_params: dict[str, str],
    dry_run: bool,
    mode: str,
    reuse_parquet: bool,
) -> None:
    params = _build_params(spec, cli_params)
    name = _resolve_string(str(spec.get("name") or "export"), params) or "export"
    output = Path(_resolve_string(spec.get("output"), params) or f"exports/{name}.xlsx")
    sheet_name = _resolve_string(spec.get("sheet_name"), params) or name[:31] or "Sheet1"
    sql = _build_sql(spec, config_dir, params)
    parquet_path = _stage_parquet_path(output, spec)

    print(f"[odps-export] {name}")
    print(sql)

    if dry_run:
        return

    rename_map = spec.get("rename") or {}
    if mode in {"all", "parquet"}:
        if reuse_parquet and parquet_path.exists():
            print(f"[odps-export] reuse parquet {parquet_path}")
        else:
            print("[odps-export] query start")
            total_rows, _ = _fetch_query_to_parquet(sql, parquet_path)
            print(f"[odps-export] staged {parquet_path} ({total_rows} rows)")

    if mode == "parquet" or output.suffix.lower() == ".parquet":
        return

    if not parquet_path.exists():
        raise FileNotFoundError(f"missing parquet stage: {parquet_path}")

    print("[odps-export] excel start")
    total_rows = _parquet_to_xlsx(parquet_path, output, sheet_name, rename_map)
    print(f"[odps-export] wrote {output} ({total_rows} rows)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export ODPS data to local Excel files")
    parser.add_argument(
        "--config",
        default=os.environ.get("ODPS_EXPORT_CONFIG", "/workspace/odps_export_config.py"),
        help="path to a project-local export config",
    )
    parser.add_argument("--dry-run", action="store_true", help="print SQL only")
    parser.add_argument(
        "--mode",
        choices=["all", "parquet", "excel"],
        default="all",
        help="run full export, only stage parquet, or only convert parquet to Excel",
    )
    parser.add_argument(
        "--reuse-parquet",
        action="store_true",
        help="skip ODPS fetch when the staged parquet file already exists",
    )
    parser.add_argument(
        "--select",
        default="",
        help="comma-separated export name(s) from EXPORTS to run",
    )
    parser.add_argument(
        "--param",
        action="append",
        default=[],
        help="template parameter override, format KEY=VALUE",
    )
    args = parser.parse_args()

    config_path = Path(args.config)
    cli_params = _parse_cli_params(args.param)
    exports = _select_exports(
        _load_config(config_path),
        [part.strip() for part in args.select.split(",") if part.strip()],
    )
    for spec in exports:
        _run_export(spec, config_path.parent, cli_params, args.dry_run, args.mode, args.reuse_parquet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
