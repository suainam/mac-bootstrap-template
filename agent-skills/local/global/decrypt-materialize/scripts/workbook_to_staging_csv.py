#!/usr/bin/env python3
"""Export selected workbook sheets to staging CSVs."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


ERROR_VALUES = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}
PLACEHOLDER_VALUES = {"NULL", "NONE", "NAN"}


def normalize(cell):
    value = cell.value
    if value is None:
        return ""
    if getattr(cell, "data_type", None) == "e":
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in PLACEHOLDER_VALUES or stripped in ERROR_VALUES:
            return ""
        return stripped
    return value


def parse_sheet_map(raw: str) -> dict[str, str]:
    data = json.loads(raw)
    if not isinstance(data, dict) or not data:
        raise ValueError("sheet map must be a non-empty JSON object")
    return {str(key): str(value) for key, value in data.items()}


def export_sheet(ws, out_path: Path) -> tuple[int, int]:
    row_count = 0
    col_count = 0
    with out_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        for row in ws.iter_rows():
            values = [normalize(cell) for cell in row]
            writer.writerow(values)
            row_count += 1
            if values:
                col_count = max(col_count, len(values))
    return row_count, col_count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--date-tag", required=True)
    parser.add_argument(
        "--sheet-map",
        required=True,
        help='JSON object, for example {"蜂窝商品数据":"hive_product_data"}',
    )
    args = parser.parse_args()

    source = Path(args.source)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    sheet_map = parse_sheet_map(args.sheet_map)

    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        summary = []
        for sheet_name, english_name in sheet_map.items():
            ws = workbook[sheet_name]
            out_path = output_dir / f"{english_name}_{args.date_tag}.csv"
            rows, cols = export_sheet(ws, out_path)
            summary.append(
                {
                    "sheet": sheet_name,
                    "file": str(out_path),
                    "rows": rows,
                    "cols": cols,
                }
            )
    finally:
        workbook.close()

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
