#!/usr/bin/env python3
"""测试工作簿物化脚本"""

import csv
import json
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


def create_test_workbook(path: Path):
    """创建测试用的 Excel 工作簿"""
    if Workbook is None:
        raise RuntimeError("需要安装 openpyxl: pip3 install openpyxl")
    wb = Workbook()

    # 工作表 1: 商品信息
    ws1 = wb.active
    ws1.title = "商品信息"
    ws1.append(["ID", "名称", "价格", "库存"])
    ws1.append([1, "苹果", 5.5, 100])
    ws1.append([2, "香蕉", 3.2, 150])
    ws1.append([3, "橙子", 4.8, 80])

    # 工作表 2: 销售数据
    ws2 = wb.create_sheet("销售数据")
    ws2.append(["日期", "商品ID", "数量", "金额"])
    ws2.append(["2026-07-01", 1, 10, 55.0])
    ws2.append(["2026-07-02", 2, 20, 64.0])
    ws2.append(["2026-07-03", 1, 5, 27.5])

    # 工作表 3: 空工作表（测试过滤）
    ws3 = wb.create_sheet("空表")

    wb.save(path)
    return wb


def verify_csv(path: Path, expected_rows: int, expected_cols: int) -> bool:
    """验证 CSV 文件"""
    if not path.exists():
        print(f"❌ 文件不存在: {path}")
        return False

    with path.open('r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)

    actual_rows = len(rows)
    actual_cols = max(len(row) for row in rows) if rows else 0

    if actual_rows != expected_rows:
        print(f"❌ 行数不匹配: {actual_rows} != {expected_rows}")
        return False

    if actual_cols != expected_cols:
        print(f"❌ 列数不匹配: {actual_cols} != {expected_cols}")
        return False

    # 检查禁止值
    forbidden = {'NULL', 'NaN', 'None'}
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell.strip() in forbidden:
                print(f"❌ 禁止值 '{cell}' 出现在 ({i}, {j})")
                return False

    print(f"✅ {path.name}: {actual_rows} 行, {actual_cols} 列")
    return True


def check_basic_export():
    """测试基本导出"""
    print("\n=== 测试 1: 基本导出 ===")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # 创建测试工作簿
        workbook = tmpdir / "测试数据.xlsx"
        create_test_workbook(workbook)

        # 运行导出
        import subprocess
        result = subprocess.run([
            sys.executable,
            "scripts/materialize.py",
            str(workbook),
            "--output-dir", str(tmpdir),
            "--date-tag", "20260713"
        ], capture_output=True, text=True, cwd=Path(__file__).parent.parent)

        if result.returncode != 0:
            print(f"❌ 导出失败: {result.stderr}")
            return False

        # 解析结果
        output = json.loads(result.stdout)
        print(f"工作簿: {output['workbook_name']}")
        print(f"加密: {output['encrypted']}")
        print(f"导出工作表数: {len(output['sheets'])}")

        # 验证输出文件
        success = True
        for sheet in output['sheets']:
            csv_path = Path(sheet['file'])
            if not verify_csv(csv_path, sheet['rows'], sheet['cols']):
                success = False

        # 检查命名规范（单下划线）
        for sheet in output['sheets']:
            filename = Path(sheet['file']).name
            if '__' in filename:
                print(f"❌ 文件名包含双下划线: {filename}")
                success = False

            expected_pattern = f"{output['workbook_name']}_"
            if not filename.startswith(expected_pattern):
                print(f"❌ 文件名不符合规范: {filename}")
                success = False

        return success


def check_sheet_mapping():
    """测试工作表映射"""
    print("\n=== 测试 2: 工作表映射 ===")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # 创建测试工作簿
        workbook = tmpdir / "产品数据.xlsx"
        create_test_workbook(workbook)

        # 使用工作表映射导出
        import subprocess
        sheet_map = json.dumps({"商品信息": "product_info", "销售数据": "sales_data"})
        result = subprocess.run([
            sys.executable,
            "scripts/materialize.py",
            str(workbook),
            "--output-dir", str(tmpdir),
            "--date-tag", "20260713",
            "--sheet-map", sheet_map
        ], capture_output=True, text=True, cwd=Path(__file__).parent.parent)

        if result.returncode != 0:
            print(f"❌ 导出失败: {result.stderr}")
            return False

        output = json.loads(result.stdout)

        # 验证映射后的文件名
        expected_files = [
            "产品数据_product_info_20260713.csv",
            "产品数据_sales_data_20260713.csv"
        ]

        success = True
        for sheet in output['sheets']:
            filename = Path(sheet['file']).name
            if filename in expected_files:
                print(f"✅ 映射文件名正确: {filename}")
            else:
                print(f"❌ 映射文件名错误: {filename}")
                success = False

        return success


def check_encryption_detection():
    """测试加密检测（模拟）"""
    print("\n=== 测试 3: 加密检测 ===")

    # 注意：实际加密的文件需要手动创建
    # 这里只测试不存在的文件

    import subprocess
    result = subprocess.run([
        sys.executable,
        "scripts/materialize.py",
        "/tmp/nonexistent_file.xlsx"
    ], capture_output=True, text=True, cwd=Path(__file__).parent.parent)

    if result.returncode != 0:
        print("✅ 正确处理了不存在的文件")
        return True
    else:
        print("❌ 应该返回错误状态码")
        return False


def main():
    """运行所有测试"""
    if Workbook is None:
        print("需要安装 openpyxl: pip3 install openpyxl")
        return 1
    print("开始测试工作簿物化脚本")
    print("=" * 50)

    tests = [
        ("基本导出", check_basic_export),
        ("工作表映射", check_sheet_mapping),
        ("加密检测", check_encryption_detection),
    ]

    results = []
    for name, test_func in tests:
        try:
            success = test_func()
            results.append((name, success))
        except Exception as e:
            print(f"❌ 测试异常: {e}")
            results.append((name, False))

    # 汇总结果
    print("\n" + "=" * 50)
    print("测试汇总:")
    passed = sum(1 for _, success in results if success)
    total = len(results)

    for name, success in results:
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"  {status}: {name}")

    print(f"\n总计: {passed}/{total} 通过")

    return 0 if passed == total else 1


if __name__ == "__main__":
    exit(main())
