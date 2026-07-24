#!/usr/bin/env python3
"""统一的工作簿物化脚本：自动检测格式和加密状态，导出为 CSV"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# 错误值和占位符定义
ERROR_VALUES = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}
PLACEHOLDER_VALUES = {"NULL", "NONE", "NAN"}


def normalize_cell(value: Any) -> str:
    """规范化单元格值"""
    if value is None:
        return ""

    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in PLACEHOLDER_VALUES or stripped in ERROR_VALUES:
            return ""
        return stripped

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value != value:  # NaN check
                return ""
            if value == int(value):
                return str(int(value))
        return str(value)

    return str(value)


def sanitize_sheet_name(name: str) -> str:
    """清理工作表名称，替换特殊字符"""
    return name.replace("/", "_").replace("\\", "_").replace(" ", "_")


def detect_encryption_openpyxl(path: Path) -> tuple[bool, str | None]:
    """检测 Excel 文件是否加密（使用 openpyxl）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        wb.close()
        return False, None
    except Exception as e:
        error_str = str(e).lower()
        if 'password' in error_str or 'encrypted' in error_str:
            return True, "Password-protected workbook"
        if 'corrupt' in error_str:
            return True, "Corrupted file"
        return True, f"Cannot open: {e}"


def export_xls(
    source: Path,
    output_dir: Path,
    date_tag: str,
    sheet_map: dict[str, str] | None = None
) -> dict[str, Any]:
    """导出 .xls 文件（使用 xlrd）"""
    try:
        import xlrd
    except ImportError:
        return {
            "error": "xlrd not installed",
            "resolution": "Run: pip3 install xlrd"
        }

    try:
        wb = xlrd.open_workbook(str(source))
    except Exception as e:
        error_str = str(e).lower()
        if 'password' in error_str or 'encrypted' in error_str:
            return {
                "source": str(source),
                "encrypted": True,
                "error": str(e),
                "resolution": "Provide password or decrypted copy"
            }
        return {
            "source": str(source),
            "error": f"Cannot open: {e}",
            "resolution": "Check if file is corrupted"
        }

    workbook_name = source.stem
    sheets_info = []

    # 确定要导出的工作表
    if sheet_map:
        target_sheets = [(wb.sheet_by_name(cn), en) for cn in sheet_map if cn in wb.sheet_names()]
    else:
        target_sheets = [(wb.sheet_by_index(i), wb.sheet_by_index(i).name) for i in range(wb.nsheets)]

    for ws, sheet_name in target_sheets:
        if ws.nrows == 0:
            sheets_info.append({
                "sheet": ws.name,
                "file": None,
                "rows": 0,
                "cols": 0,
                "skipped": True,
                "reason": "empty sheet, no CSV written"
            })
            continue

        safe_name = sanitize_sheet_name(sheet_name) if not sheet_map else sheet_name
        output_path = output_dir / f"{workbook_name}_{safe_name}_{date_tag}.csv"

        # 导出 CSV
        with output_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row_idx in range(ws.nrows):
                values = [normalize_cell(ws.cell_value(row_idx, col_idx)) for col_idx in range(ws.ncols)]
                writer.writerow(values)

        sheets_info.append({
            "sheet": ws.name,
            "file": str(output_path),
            "rows": ws.nrows,
            "cols": ws.ncols,
            "verified": output_path.exists()
        })

    return {
        "source": str(source),
        "workbook_name": workbook_name,
        "encrypted": False,
        "date_tag": date_tag,
        "sheets": sheets_info
    }


def export_xlsx(
    source: Path,
    output_dir: Path,
    date_tag: str,
    sheet_map: dict[str, str] | None = None
) -> dict[str, Any]:
    """导出 .xlsx 文件（使用 openpyxl）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "error": "openpyxl not installed",
            "resolution": "Run: pip3 install openpyxl"
        }

    # 检测加密
    encrypted, error_msg = detect_encryption_openpyxl(source)
    if encrypted:
        return {
            "source": str(source),
            "encrypted": True,
            "error": error_msg,
            "resolution": "Provide password or decrypted copy"
        }

    # 加载工作簿
    wb = load_workbook(source, read_only=True, data_only=True)
    workbook_name = source.stem
    sheets_info = []

    try:
        # 确定要导出的工作表
        if sheet_map:
            target_sheets = [(wb[cn], en) for cn, en in sheet_map.items() if cn in wb.sheetnames]
        else:
            target_sheets = [(ws, ws.title) for ws in wb.worksheets]

        for ws, sheet_name in target_sheets:
            # 构建输出文件名：workbook_sheet_date.csv
            safe_name = sanitize_sheet_name(sheet_name) if not sheet_map else sheet_name
            output_path = output_dir / f"{workbook_name}_{safe_name}_{date_tag}.csv"

            # 导出 CSV（跳过空行）
            row_count = 0
            col_count = 0
            with output_path.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows():
                    values = [normalize_cell(cell.value) for cell in row]
                    # 跳过完全空的行
                    if any(v for v in values):
                        writer.writerow(values)
                        row_count += 1
                        col_count = max(col_count, len(values))

            if row_count == 0:
                output_path.unlink(missing_ok=True)
                sheets_info.append({
                    "sheet": ws.title,
                    "file": None,
                    "rows": 0,
                    "cols": 0,
                    "skipped": True,
                    "reason": "empty sheet, no CSV written"
                })
                continue

            sheets_info.append({
                "sheet": ws.title,
                "file": str(output_path),
                "rows": row_count,
                "cols": col_count,
                "verified": output_path.exists()
            })

    finally:
        wb.close()

    return {
        "source": str(source),
        "workbook_name": workbook_name,
        "encrypted": False,
        "date_tag": date_tag,
        "sheets": sheets_info
    }


def export_numbers(
    source: Path,
    output_dir: Path,
    date_tag: str,
    sheet_map: dict[str, str] | None = None
) -> dict[str, Any]:
    """导出 Numbers 文件"""
    try:
        import numbers_parser
    except ImportError:
        return {
            "error": "numbers-parser not installed",
            "resolution": 'Run: pip3 install "numbers-parser==4.18.5"'
        }

    try:
        doc = numbers_parser.Document(str(source))
    except Exception as e:
        return {
            "source": str(source),
            "encrypted": True,
            "error": f"Cannot open: {e}",
            "resolution": "Check if file is corrupted or encrypted"
        }

    workbook_name = source.stem
    sheets_info = []

    # 获取有数据的工作表
    sheets_with_data = [
        sheet for sheet in doc.sheets
        if sheet.tables and _has_data_numbers(sheet.tables[0])
    ]

    if not sheets_with_data:
        return {
            "source": str(source),
            "error": "No sheets with data found",
            "sheets": []
        }

    # 确定要导出的工作表
    if sheet_map:
        target_sheets = [(s, sheet_map[s.name]) for s in sheets_with_data if s.name in sheet_map]
    else:
        target_sheets = [(s, s.name) for s in sheets_with_data]

    for sheet, sheet_name in target_sheets:
        safe_name = sanitize_sheet_name(sheet_name) if not sheet_map else sheet_name
        output_path = output_dir / f"{workbook_name}_{safe_name}_{date_tag}.csv"

        # 导出
        row_count = 0
        col_count = 0
        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in sheet.tables[0].iter_rows():
                values = [_cell_to_str(cell) for cell in row]
                writer.writerow(values)
                row_count += 1
                if values:
                    col_count = max(col_count, len(values))

        sheets_info.append({
            "sheet": sheet.name,
            "file": str(output_path),
            "rows": row_count,
            "cols": col_count,
            "verified": output_path.exists()
        })

    return {
        "source": str(source),
        "workbook_name": workbook_name,
        "encrypted": False,
        "date_tag": date_tag,
        "sheets": sheets_info
    }


def _cell_to_str(cell) -> str:
    """Numbers cell 转字符串"""
    try:
        value = cell.value
    except Exception:
        return ""
    return normalize_cell(value)


def _has_data_numbers(table) -> bool:
    """检查 Numbers 表格是否有数据"""
    check_rows = min(3, table.num_rows)
    for row in table.iter_rows(min_row=0, max_row=check_rows - 1):
        for cell in row:
            if _cell_to_str(cell):
                return True
    return False


def parse_sheet_map(raw: str) -> dict[str, str]:
    """解析工作表映射 JSON"""
    data = json.loads(raw)
    if not isinstance(data, dict) or not data:
        raise ValueError("sheet-map must be a non-empty JSON object")
    return {str(key): str(value) for key, value in data.items()}


def resolve_output_dir(explicit_dir: Path | None, source: Path) -> Path:
    """按优先级解析输出目录：显式指定 > ../02_working_data/ > ./decrypted/ > 当前目录"""
    if explicit_dir:
        if not explicit_dir.exists():
            print(f"ERROR: Specified output directory does not exist: {explicit_dir}", file=sys.stderr)
            sys.exit(1)
        return explicit_dir

    # 优先级 1: ../02_working_data/
    candidate = source.parent.parent / "02_working_data"
    if candidate.exists():
        return candidate

    # 优先级 2: ./decrypted/
    candidate = Path.cwd() / "decrypted"
    if candidate.exists():
        return candidate

    # 优先级 3: 当前目录
    return Path.cwd()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="统一的工作簿物化脚本：自动检测格式和加密状态"
    )
    parser.add_argument("source", type=Path, help="源工作簿文件路径")
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="输出目录（显式指定则必须存在；否则按 ../02_working_data/ > ./decrypted/ > . 选择）"
    )
    parser.add_argument(
        "--date-tag",
        help="日期标签（YYYYMMDD），默认：今天"
    )
    parser.add_argument(
        "--sheet-map",
        help='工作表映射 JSON，例如 \'{"中文名":"english_name"}\''
    )

    args = parser.parse_args()

    # 验证源文件
    if not args.source.exists():
        print(f"ERROR: File not found: {args.source}", file=sys.stderr)
        return 1

    # 解析输出目录
    output_dir = resolve_output_dir(args.output_dir, args.source)

    # 设置日期标签
    date_tag = args.date_tag or datetime.now().strftime("%Y%m%d")

    # 解析工作表映射
    sheet_map = parse_sheet_map(args.sheet_map) if args.sheet_map else None

    # 根据扩展名选择导出器
    ext = args.source.suffix.lower()
    if ext == ".numbers":
        result = export_numbers(args.source, output_dir, date_tag, sheet_map)
    elif ext == ".xlsx":
        result = export_xlsx(args.source, output_dir, date_tag, sheet_map)
    elif ext == ".xls":
        result = export_xls(args.source, output_dir, date_tag, sheet_map)
    else:
        print(f"ERROR: Unsupported format: {ext}", file=sys.stderr)
        print("Supported: .numbers, .xlsx, .xls", file=sys.stderr)
        return 1

    # 输出结果
    print(json.dumps(result, ensure_ascii=False, indent=2))

    # 如果有错误，返回非零状态码
    if "error" in result:
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
