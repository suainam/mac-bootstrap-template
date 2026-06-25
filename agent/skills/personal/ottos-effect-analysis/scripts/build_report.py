#!/usr/bin/env python3
"""
效果分析 xlsx 报告生成 —— 把固定的 openpyxl 样式样板封装掉。

使用者只准备一个描述报告内容的 JSON（结论 + 各 sheet 的表格数据），
脚本负责出一份带统一样式的多 sheet xlsx：标题合并单元格、表头蓝底白字、
正向绿字 / 负向红字、结论区浅黄底、列宽自适应。

配置里 rows 的单元格可以直接写 Excel 公式（以 "=" 开头），脚本原样写入，
由 Excel/LibreOffice 打开时计算——符合"用公式而非硬编码"的要求。

用法：
    python3 build_report.py config.json [output.xlsx]
配置结构见 references/xlsx-report.md，或运行 `python3 build_report.py` 看示例。

列索引（pct_cols / signed_cols / num_cols）一律用 1 起的列号（A=1, B=2 …）。
"""
import sys
import json
import subprocess


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=False)


EXAMPLE = {
    "strategy_name": "门店自动陈列",
    "date_range": "2026-03 ~ 2026-05",
    "version": "v0.x 草稿",
    "summary": {
        "one_line": "自动陈列净效果约 +4pp 动销率，归因可信度中，建议分层推广。",
        "key_metrics": [["动销率净效果", "+4pp"], ["米效净效果", "+5.9%"], ["归因方法", "同比+区域+DID"]],
        "confidence": "中",
        "confidence_basis": "同比与区域对比方向一致；DID 仅 5 对门店，t 检验不显著但 5/5 同向。",
        "next_step": "在条件较好的营运区先分层推广，3 个月后复盘是否衰减。",
        "risks": "缺货率同比上升 1.2pp，需持续监控护栏指标。",
    },
    "sheets": [
        {
            "name": "同比趋势",
            "headers": ["指标", "去年同期", "今年", "同比变化"],
            "rows": [
                ["动销率", 0.42, 0.46, "=C2-B2"],
                ["米效(元/米/天)", 85, 90, "=C3-B3"],
            ],
            "pct_cols": [2, 3],
            "signed_cols": [4],
        }
    ],
}


def build(config, out_path):
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    data_font = Font(name="Arial", size=10)
    label_font = Font(name="Arial", size=10, bold=True)
    pos_font = Font(name="Arial", size=10, color="008000")
    neg_font = Font(name="Arial", size=10, color="FF0000")
    conclusion_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center")

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            # 中文按 2 个宽度近似
            extra = max((sum(1 for ch in str(c.value) if ord(ch) > 127) for c in col if c.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + extra + 2, 40)

    wb = Workbook()

    # Sheet1 总览与结论
    s = config.get("summary", {})
    ws = wb.active
    ws.title = "总览与结论"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{config.get('strategy_name','')} 效果分析报告　{config.get('date_range','')}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    rows = [
        ("编制", "数智中台部 · 数据运营组"),
        ("版本", config.get("version", "v0.x 草稿")),
        ("一句话结论", s.get("one_line", "")),
        ("归因可信度", s.get("confidence", "")),
        ("可信度依据", s.get("confidence_basis", "")),
        ("下一步建议", s.get("next_step", "")),
        ("风险/代价", s.get("risks", "")),
    ]
    r = 3
    for label, val in rows:
        ws.cell(r, 1, label).font = label_font
        ws.cell(r, 1).fill = conclusion_fill
        c = ws.cell(r, 2, val)
        c.font = data_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1
    if s.get("key_metrics"):
        ws.cell(r, 1, "核心数字").font = label_font
        r += 1
        for km in s["key_metrics"]:
            ws.cell(r, 1, km[0]).font = data_font
            ws.cell(r, 2, km[1]).font = data_font
            r += 1
    ws.column_dimensions["A"].width = 16
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 22

    # 数据 sheet
    for sh in config.get("sheets", []):
        ws = wb.create_sheet(sh["name"][:31])
        headers = sh.get("headers", [])
        for j, h in enumerate(headers, 1):
            c = ws.cell(1, j, h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        pct = set(sh.get("pct_cols", []))
        signed = set(sh.get("signed_cols", []))
        num = set(sh.get("num_cols", []))
        for i, row in enumerate(sh.get("rows", []), 2):
            for j, val in enumerate(row, 1):
                c = ws.cell(i, j, val)
                c.font = data_font
                if j in pct:
                    c.number_format = "0.0%"
                elif j in num:
                    c.number_format = "#,##0"
        # 正负上色用条件格式：按单元格实际值（含公式计算结果）染色——公式列也能正确区分正负
        nrows = len(sh.get("rows", []))
        if nrows and signed:
            for col in signed:
                cl = get_column_letter(col)
                rng = f"{cl}2:{cl}{nrows + 1}"
                ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=pos_font))
                ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=neg_font))
        if sh.get("note"):
            nr = len(sh.get("rows", [])) + 2
            ws.cell(nr, 1, sh["note"]).font = Font(name="Arial", size=9, italic=True, color="808080")
        autosize(ws)

    wb.save(out_path)
    return out_path


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("用法：python3 build_report.py config.json [output.xlsx]", file=sys.stderr)
        print("\n配置 JSON 示例：", file=sys.stderr)
        print(json.dumps(EXAMPLE, ensure_ascii=False, indent=2), file=sys.stderr)
        sys.exit(1)
    with open(args[0], encoding="utf-8") as f:
        config = json.load(f)
    out = args[1] if len(args) > 1 else f"{config.get('strategy_name','策略')}-效果分析报告.xlsx"
    path = build(config, out)
    print(f"已生成：{path}")
    print("提示：含公式时用 LibreOffice/Excel 打开会自动计算；如需校验公式可用 xlsx skill 的 recalc.py。")


if __name__ == "__main__":
    main()
